import os
import openpyxl
import xlrd
import sys
import threading
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QLineEdit, QTextBrowser, QVBoxLayout, QFileDialog, QProgressBar
from PyQt5.QtCore import QObject, pyqtSignal, pyqtSlot, Qt

class ExcelSearchApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.search_thread = None

    def initUI(self):
        self.setWindowTitle('Excel Search Tool')
        self.setGeometry(100, 100, 600, 400)

        self.folder_label = QLabel('Select Folder:')
        self.folder_button = QPushButton('Browse')
        self.folder_button.clicked.connect(self.browse_folder)
        self.folder_line_edit = QLineEdit()

        self.search_label = QLabel('Enter Word to Search:')
        self.search_line_edit = QLineEdit()

        self.search_button = QPushButton('Search')
        self.search_button.clicked.connect(self.start_search)

        self.result_text_browser = QTextBrowser()
        self.progress_bar = QProgressBar()

        layout = QVBoxLayout()
        layout.addWidget(self.folder_label)
        layout.addWidget(self.folder_button)
        layout.addWidget(self.folder_line_edit)
        layout.addWidget(self.search_label)
        layout.addWidget(self.search_line_edit)
        layout.addWidget(self.search_button)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.result_text_browser)

        self.setLayout(layout)

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, 'Select Folder')
        if folder:
            self.folder_line_edit.setText(folder)

    def start_search(self):
        folder_path = self.folder_line_edit.text()
        search_word = self.search_line_edit.text()

        # Check if a search is already running, if so, stop it
        if self.search_thread and self.search_thread.is_alive():
            self.search_thread.stop_search()

        # Clear the previous search results and reset the progress bar
        self.result_text_browser.clear()
        self.progress_bar.setValue(0)

        # Create a new search thread
        self.search_thread = SearchThread(folder_path, search_word)
        self.search_thread.search_complete.connect(self.display_results)
        self.search_thread.search_progress.connect(self.update_progress)
        self.search_thread.start()

    @pyqtSlot(list)
    def display_results(self, matching_info):
        if matching_info:
            result_text = "Matching occurrences:\n"
            for match in matching_info:
                result_text += f"File: {match['File']}, Sheet: {match['Sheet']}, Cell: {match['Cell']}\n"
            self.result_text_browser.setPlainText(result_text)
        else:
            self.result_text_browser.setPlainText("No matching occurrences found.")

    @pyqtSlot(int)
    def update_progress(self, value):
        self.progress_bar.setValue(value)

class SearchThread(QObject, threading.Thread):
    search_complete = pyqtSignal(list)
    search_progress = pyqtSignal(int)

    def __init__(self, folder_path, search_word):
        super().__init__()
        self.folder_path = folder_path
        self.search_word = search_word
        self.is_running = True

    def run(self):
        matching_info = []

        for root, _, files in os.walk(self.folder_path):
            if not self.is_running:
                break

            total_files = len(files)
            processed_files = 0

            for filename in files:
                if not self.is_running:
                    break

                if filename.endswith('.xlsx'):
                    file_path = os.path.join(root, filename)
                    workbook = openpyxl.load_workbook(file_path)
                    for sheet_name in workbook.sheetnames:
                        if not self.is_running:
                            break

                        sheet = workbook[sheet_name]
                        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            for column_number, cell_value in enumerate(row, start=1):
                                if not self.is_running:
                                    break

                                if self.search_word in str(cell_value):
                                    matching_info.append({
                                        "File": file_path,
                                        "Sheet": sheet.title,
                                        "Cell": f"{openpyxl.utils.get_column_letter(column_number)}{row_number}",
                                    })

                elif filename.endswith('.xls'):
                    file_path = os.path.join(root, filename)
                    xls_workbook = xlrd.open_workbook(file_path)
                    for sheet in xls_workbook.sheets():
                        if not self.is_running:
                            break

                        for row_number in range(sheet.nrows):
                            for column_number in range(sheet.ncols):
                                if not self.is_running:
                                    break

                                cell_value = sheet.cell_value(row_number, column_number)
                                if self.search_word in str(cell_value):
                                    matching_info.append({
                                        "File": file_path,
                                        "Sheet": sheet.name,
                                        "Cell": f"{xlrd.colname(column_number)}{row_number + 1}",
                                    })

                processed_files += 1
                progress = int((processed_files / total_files) * 100)
                self.search_progress.emit(progress)

        # Emit the signal with the search results
        self.search_complete.emit(matching_info)

    def stop_search(self):
        self.is_running = False

def main():
    app = QApplication(sys.argv)
    window = ExcelSearchApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
